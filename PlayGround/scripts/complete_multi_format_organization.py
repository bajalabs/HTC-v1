#!/usr/bin/env python3
"""
Complete Multi-Format HTS Organization Script
Finds and matches CSV, JSON, and Excel files by chapter, then organizes them together.
This handles the case where each format was downloaded to separate timestamp directories.
"""

import os
import csv
import json
import shutil
from pathlib import Path
from collections import defaultdict
import re
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not available. Install with: pip install openpyxl")

# Chapter names mapping
CHAPTER_NAMES = {
    1: "live-animals", 2: "meat-and-edible-meat-offal", 3: "fish-and-crustaceans",
    4: "dairy-produce-birds-eggs", 5: "products-of-animal-origin", 
    7: "edible-vegetables-and-roots", 8: "edible-fruit-and-nuts",
    9: "coffee-tea-mate-spices", 10: "cereals", 11: "milling-industry-products",
    12: "oil-seeds-oleaginous-fruits", 13: "lac-gums-resins-vegetable-extracts",
    14: "vegetable-plaiting-materials", 15: "animal-vegetable-fats-oils",
    16: "meat-fish-preparations", 17: "sugars-sugar-confectionery",
    18: "cocoa-cocoa-preparations", 19: "cereal-flour-starch-preparations",
    20: "vegetable-fruit-preparations", 21: "miscellaneous-edible-preparations",
    22: "beverages-spirits-vinegar", 23: "food-industry-residues-animal-feed",
    24: "tobacco-manufactured-tobacco", 25: "salt-sulfur-earths-stone",
    26: "ores-slag-ash", 27: "mineral-fuels-oils", 28: "inorganic-chemicals",
    29: "organic-chemicals", 30: "pharmaceutical-products", 31: "fertilizers",
    32: "tanning-dyeing-extracts", 33: "essential-oils-perfumery",
    34: "soap-organic-surface-agents", 35: "albuminoidal-substances",
    36: "explosives-pyrotechnics", 37: "photographic-cinematographic-goods",
    38: "miscellaneous-chemical-products", 39: "plastics-articles-thereof",
    40: "rubber-articles-thereof", 41: "raw-hides-skins-leather",
    42: "leather-articles-saddlery", 43: "furskins-artificial-fur",
    44: "wood-articles-wood-charcoal", 45: "cork-articles-cork",
    46: "straw-plaiting-materials", 47: "wood-pulp-fibrous-material",
    48: "paper-paperboard-articles", 49: "printed-books-newspapers",
    50: "silk", 51: "wool-animal-hair", 52: "cotton",
    53: "vegetable-textile-fibers", 54: "man-made-filaments",
    55: "man-made-staple-fibers", 56: "wadding-felt-nonwovens",
    57: "carpets-textile-floor-coverings", 58: "special-woven-fabrics",
    59: "impregnated-coated-textile-fabrics", 60: "knitted-crocheted-fabrics",
    61: "knitted-crocheted-apparel", 62: "not-knitted-crocheted-apparel",
    63: "textile-articles-worn-clothing", 64: "footwear-gaiters",
    65: "headgear-parts-thereof", 66: "umbrellas-walking-sticks",
    67: "prepared-feathers-artificial-flowers", 68: "stone-plaster-cement-articles",
    69: "ceramic-products", 70: "glass-glassware",
    71: "pearls-precious-stones-metals", 72: "iron-steel",
    73: "iron-steel-articles", 74: "copper-articles-thereof",
    75: "nickel-articles-thereof", 76: "aluminum-articles-thereof",
    78: "lead-articles-thereof", 79: "zinc-articles-thereof",
    80: "tin-articles-thereof", 81: "base-metals-cermets",
    82: "tools-implements-cutlery", 83: "miscellaneous-base-metal-articles",
    84: "nuclear-reactors-boilers-machinery", 85: "electrical-machinery-equipment",
    86: "railway-locomotives-rolling-stock", 87: "vehicles-parts-accessories",
    88: "aircraft-spacecraft-parts", 89: "ships-boats-floating-structures",
    90: "optical-photographic-instruments", 91: "clocks-watches-parts",
    92: "musical-instruments-parts", 93: "arms-ammunition-parts",
    94: "furniture-bedding-lamps", 95: "toys-games-sports",
    96: "miscellaneous-manufactured-articles", 97: "works-of-art-collectors-pieces",
    98: "special-classification-provisions", 99: "temporary-legislation-modifications"
}

def identify_chapter_from_csv(csv_file):
    """Identify chapter from CSV file."""
    try:
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader)
            first_row = next(reader)
            hts_number = first_row[0].strip('"').strip()
            
            if len(hts_number) >= 2 and hts_number.isdigit():
                chapter_num = int(hts_number[:2])
                if 1 <= chapter_num <= 99:
                    return chapter_num
    except Exception as e:
        print(f"  Error reading CSV {csv_file}: {e}")
    return None

def identify_chapter_from_json(json_file):
    """Identify chapter from JSON file."""
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
            if isinstance(data, list) and len(data) > 0:
                first_data = data[0]
                
                hts_number = None
                if isinstance(first_data, dict):
                    for field_name in ['htsno', 'HTS Number', 'HTS_Number', 'hts_number', 'HTS', 'hts']:
                        if field_name in first_data:
                            hts_number = str(first_data[field_name]).strip('"').strip()
                            break
                elif isinstance(first_data, list) and len(first_data) > 0:
                    hts_number = str(first_data[0]).strip('"').strip()
                
                if hts_number and len(hts_number) >= 2 and hts_number.isdigit():
                    chapter_num = int(hts_number[:2])
                    if 1 <= chapter_num <= 99:
                        return chapter_num
    except Exception as e:
        print(f"  Error reading JSON {json_file}: {e}")
    return None

def identify_chapter_from_excel(excel_file):
    """Identify chapter from Excel file."""
    if not OPENPYXL_AVAILABLE:
        return None
    
    try:
        workbook = load_workbook(excel_file, read_only=True)
        worksheet = workbook.active
        
        for row in worksheet.iter_rows(min_row=2, max_row=3, values_only=True):
            if row and row[0]:
                hts_number = str(row[0]).strip('"').strip()
                
                if len(hts_number) >= 2:
                    try:
                        chapter_num = int(hts_number[:2])
                        if 1 <= chapter_num <= 99:
                            workbook.close()
                            return chapter_num
                    except ValueError:
                        continue
        
        workbook.close()
    except Exception as e:
        print(f"  Error reading Excel {excel_file}: {e}")
    return None

def scan_all_files(playwright_dir):
    """Scan all files and categorize by chapter and format."""
    base_path = Path(playwright_dir)
    timestamp_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}T\d{2}-\d{2}-\d{2}\.\d{3}Z$')
    
    # Dictionary to store files by chapter: {chapter_num: {'csv': path, 'json': path, 'xlsx': path}}
    chapter_files = defaultdict(lambda: {'csv': None, 'json': None, 'xlsx': None})
    unidentified_files = []
    
    print("🔍 Scanning all timestamp directories...")
    
    total_dirs = 0
    processed_dirs = 0
    
    for item in base_path.iterdir():
        if item.is_dir() and timestamp_pattern.match(item.name):
            total_dirs += 1
            
            # Check what files are in this directory
            csv_file = item / "htsdata.csv"
            json_file = item / "htsdata.json"
            xlsx_file = item / "htsdata.xlsx"
            
            chapter_num = None
            file_type = None
            
            # Identify chapter and file type
            if csv_file.exists():
                chapter_num = identify_chapter_from_csv(csv_file)
                file_type = 'csv'
                file_path = csv_file
            elif json_file.exists():
                chapter_num = identify_chapter_from_json(json_file)
                file_type = 'json'
                file_path = json_file
            elif xlsx_file.exists():
                chapter_num = identify_chapter_from_excel(xlsx_file)
                file_type = 'xlsx'
                file_path = xlsx_file
            
            if chapter_num and file_type:
                chapter_name = CHAPTER_NAMES.get(chapter_num, f"chapter-{chapter_num}")
                print(f"  📁 {item.name}: Chapter {chapter_num:02d} ({chapter_name}) - {file_type.upper()}")
                
                # Store the file path for this chapter and format
                if chapter_files[chapter_num][file_type] is None:
                    chapter_files[chapter_num][file_type] = file_path
                    processed_dirs += 1
                else:
                    print(f"    ⚠️  Duplicate {file_type.upper()} found for Chapter {chapter_num}, keeping first one")
            else:
                print(f"  ❓ {item.name}: Could not identify chapter")
                unidentified_files.append(item)
    
    print(f"\n📊 Scan Summary:")
    print(f"  Total directories: {total_dirs}")
    print(f"  Successfully processed: {processed_dirs}")
    print(f"  Unidentified: {len(unidentified_files)}")
    print(f"  Unique chapters found: {len(chapter_files)}")
    
    return dict(chapter_files), unidentified_files

def organize_complete_chapters(chapter_files, chapters_dir):
    """Organize all files into chapter directories with complete format coverage."""
    chapters_path = Path(chapters_dir)
    chapters_path.mkdir(parents=True, exist_ok=True)
    
    print(f"\n📁 Organizing files into {chapters_dir}...")
    
    successful_copies = 0
    failed_copies = []
    complete_chapters = 0
    
    for chapter_num in sorted(chapter_files.keys()):
        chapter_name = CHAPTER_NAMES.get(chapter_num, f"chapter-{chapter_num}")
        chapter_dir = chapters_path / f"chapter-{chapter_num:02d}"
        chapter_dir.mkdir(exist_ok=True)
        
        files = chapter_files[chapter_num]
        csv_file = files['csv']
        json_file = files['json']
        xlsx_file = files['xlsx']
        
        # Count available formats
        available_formats = sum(1 for f in [csv_file, json_file, xlsx_file] if f is not None)
        
        print(f"\n📂 Chapter {chapter_num:02d}: {chapter_name} ({available_formats}/3 formats)")
        
        # Copy each available file
        for file_type, source_path in [('csv', csv_file), ('json', json_file), ('xlsx', xlsx_file)]:
            if source_path and source_path.exists():
                target_filename = f"chapter-{chapter_num:02d}-{chapter_name}.{file_type}"
                target_path = chapter_dir / target_filename
                
                try:
                    # Only copy if target doesn't exist or source is newer
                    if not target_path.exists() or source_path.stat().st_mtime > target_path.stat().st_mtime:
                        shutil.copy2(source_path, target_path)
                        print(f"  ✅ {file_type.upper()}: {target_filename}")
                        successful_copies += 1
                    else:
                        print(f"  ⏭️  {file_type.upper()}: {target_filename} (already exists)")
                except Exception as e:
                    error_msg = f"Failed to copy {source_path} to {target_path}: {e}"
                    print(f"  ❌ {file_type.upper()}: {error_msg}")
                    failed_copies.append(error_msg)
            else:
                print(f"  ⚪ {file_type.upper()}: not available")
        
        # Check if chapter is complete (has all 3 formats)
        if all(files[ft] is not None for ft in ['csv', 'json', 'xlsx']):
            complete_chapters += 1
    
    return successful_copies, failed_copies, complete_chapters

def generate_comprehensive_report(chapter_files, unidentified_files, successful_copies, failed_copies, complete_chapters, output_dir):
    """Generate comprehensive organization report."""
    report_path = Path(output_dir) / "analysis" / "comprehensive_organization_report.md"
    report_path.parent.mkdir(exist_ok=True)
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("# Comprehensive HTS Organization Report\n\n")
        
        # Summary statistics
        total_chapters = len(chapter_files)
        csv_count = sum(1 for files in chapter_files.values() if files['csv'] is not None)
        json_count = sum(1 for files in chapter_files.values() if files['json'] is not None)
        xlsx_count = sum(1 for files in chapter_files.values() if files['xlsx'] is not None)
        
        f.write("## Summary Statistics\n\n")
        f.write(f"- **Total Chapters Found**: {total_chapters}\n")
        f.write(f"- **Complete Chapters** (all 3 formats): {complete_chapters}\n")
        f.write(f"- **CSV Files**: {csv_count}\n")
        f.write(f"- **JSON Files**: {json_count}\n")
        f.write(f"- **Excel Files**: {xlsx_count}\n")
        f.write(f"- **Total Files Copied**: {successful_copies}\n")
        f.write(f"- **Failed Copies**: {len(failed_copies)}\n")
        f.write(f"- **Unidentified Directories**: {len(unidentified_files)}\n\n")
        
        # Chapter details table
        f.write("## Chapter Details\n\n")
        f.write("| Chapter | Name | CSV | JSON | Excel | Complete |\n")
        f.write("|---------|------|-----|------|-------|----------|\n")
        
        for chapter_num in sorted(chapter_files.keys()):
            chapter_name = CHAPTER_NAMES.get(chapter_num, f"chapter-{chapter_num}")
            files = chapter_files[chapter_num]
            
            csv_status = "✅" if files['csv'] else "❌"
            json_status = "✅" if files['json'] else "❌"
            xlsx_status = "✅" if files['xlsx'] else "❌"
            complete_status = "✅" if all(files[ft] for ft in ['csv', 'json', 'xlsx']) else "❌"
            
            f.write(f"| {chapter_num:02d} | {chapter_name} | {csv_status} | {json_status} | {xlsx_status} | {complete_status} |\n")
        
        # Missing chapters
        expected_chapters = set(range(1, 100)) - {6, 77}  # Known missing chapters
        found_chapters = set(chapter_files.keys())
        missing_chapters = expected_chapters - found_chapters
        
        if missing_chapters:
            f.write(f"\n## Missing Chapters\n\n")
            f.write(f"Chapters not found: {', '.join(map(str, sorted(missing_chapters)))}\n\n")
        
        # Failed copies
        if failed_copies:
            f.write(f"\n## Failed Copy Operations\n\n")
            for error in failed_copies:
                f.write(f"- {error}\n")
        
        # Unidentified files
        if unidentified_files:
            f.write(f"\n## Unidentified Directories\n\n")
            for file_info in unidentified_files[:10]:  # Show first 10
                f.write(f"- {file_info.name}\n")
            if len(unidentified_files) > 10:
                f.write(f"- ... and {len(unidentified_files) - 10} more\n")
    
    print(f"\n📋 Comprehensive report saved to {report_path}")

def main():
    """Main function."""
    playwright_dir = "../playwright-mcp-output11"
    chapters_dir = "../chapters"
    output_dir = "../"
    
    print("🎯 Comprehensive Multi-Format HTS Organization")
    print("=" * 60)
    
    # Step 1: Scan all files
    chapter_files, unidentified_files = scan_all_files(playwright_dir)
    
    # Step 2: Organize files
    successful_copies, failed_copies, complete_chapters = organize_complete_chapters(
        chapter_files, chapters_dir
    )
    
    # Step 3: Generate report
    generate_comprehensive_report(
        chapter_files, unidentified_files, successful_copies, failed_copies, 
        complete_chapters, output_dir
    )
    
    # Final summary
    print(f"\n🎉 COMPREHENSIVE ORGANIZATION COMPLETE!")
    print(f"=" * 60)
    print(f"📊 {len(chapter_files)} chapters processed")
    print(f"✅ {complete_chapters} chapters have all 3 formats")
    print(f"📄 {successful_copies} files successfully organized")
    
    if failed_copies:
        print(f"⚠️  {len(failed_copies)} copy operations failed")
    if unidentified_files:
        print(f"❓ {len(unidentified_files)} directories could not be identified")
    
    print(f"\n📁 All files organized in: {chapters_dir}/")
    print(f"📋 Detailed report: analysis/comprehensive_organization_report.md")

if __name__ == "__main__":
    main()
