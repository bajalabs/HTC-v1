#!/usr/bin/env python3
"""
Excel Files Organization Script
Finds and organizes all Excel (.xlsx) HTS files by chapter.
"""

import os
import csv
import shutil
from pathlib import Path
from collections import defaultdict
import re
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not available. Install with: pip install openpyxl")

# Chapter names mapping
CHAPTER_NAMES = {
    1: "live-animals", 2: "meat-and-edible-meat-offal", 3: "fish-and-crustaceans",
    4: "dairy-produce-birds-eggs", 5: "products-of-animal-origin", 
    7: "edible-vegetables-and-roots", 8: "edible-fruit-and-nuts",
    9: "coffee-tea-mate-spices", 10: "cereals", 11: "milling-industry-products",
    12: "oil-seeds-oleaginous-fruits", 13: "lac-gums-resins-vegetable-extracts",
    14: "vegetable-plaiting-materials", 15: "animal-vegetable-fats-oils",
    16: "meat-fish-preparations", 17: "sugars-sugar-confectionery",
    18: "cocoa-cocoa-preparations", 19: "cereal-flour-starch-preparations",
    20: "vegetable-fruit-preparations", 21: "miscellaneous-edible-preparations",
    22: "beverages-spirits-vinegar", 23: "food-industry-residues-animal-feed",
    24: "tobacco-manufactured-tobacco", 25: "salt-sulfur-earths-stone",
    26: "ores-slag-ash", 27: "mineral-fuels-oils", 28: "inorganic-chemicals",
    29: "organic-chemicals", 30: "pharmaceutical-products", 31: "fertilizers",
    32: "tanning-dyeing-extracts", 33: "essential-oils-perfumery",
    34: "soap-organic-surface-agents", 35: "albuminoidal-substances",
    36: "explosives-pyrotechnics", 37: "photographic-cinematographic-goods",
    38: "miscellaneous-chemical-products", 39: "plastics-articles-thereof",
    40: "rubber-articles-thereof", 41: "raw-hides-skins-leather",
    42: "leather-articles-saddlery", 43: "furskins-artificial-fur",
    44: "wood-articles-wood-charcoal", 45: "cork-articles-cork",
    46: "straw-plaiting-materials", 47: "wood-pulp-fibrous-material",
    48: "paper-paperboard-articles", 49: "printed-books-newspapers",
    50: "silk", 51: "wool-animal-hair", 52: "cotton",
    53: "vegetable-textile-fibers", 54: "man-made-filaments",
    55: "man-made-staple-fibers", 56: "wadding-felt-nonwovens",
    57: "carpets-textile-floor-coverings", 58: "special-woven-fabrics",
    59: "impregnated-coated-textile-fabrics", 60: "knitted-crocheted-fabrics",
    61: "knitted-crocheted-apparel", 62: "not-knitted-crocheted-apparel",
    63: "textile-articles-worn-clothing", 64: "footwear-gaiters",
    65: "headgear-parts-thereof", 66: "umbrellas-walking-sticks",
    67: "prepared-feathers-artificial-flowers", 68: "stone-plaster-cement-articles",
    69: "ceramic-products", 70: "glass-glassware",
    71: "pearls-precious-stones-metals", 72: "iron-steel",
    73: "iron-steel-articles", 74: "copper-articles-thereof",
    75: "nickel-articles-thereof", 76: "aluminum-articles-thereof",
    78: "lead-articles-thereof", 79: "zinc-articles-thereof",
    80: "tin-articles-thereof", 81: "base-metals-cermets",
    82: "tools-implements-cutlery", 83: "miscellaneous-base-metal-articles",
    84: "nuclear-reactors-boilers-machinery", 85: "electrical-machinery-equipment",
    86: "railway-locomotives-rolling-stock", 87: "vehicles-parts-accessories",
    88: "aircraft-spacecraft-parts", 89: "ships-boats-floating-structures",
    90: "optical-photographic-instruments", 91: "clocks-watches-parts",
    92: "musical-instruments-parts", 93: "arms-ammunition-parts",
    94: "furniture-bedding-lamps", 95: "toys-games-sports",
    96: "miscellaneous-manufactured-articles", 97: "works-of-art-collectors-pieces",
    98: "special-classification-provisions", 99: "temporary-legislation-modifications"
}

def find_excel_files(playwright_dir):
    """Find all Excel (.xlsx) files in timestamp directories."""
    base_path = Path(playwright_dir)
    timestamp_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}T\d{2}-\d{2}-\d{2}\.\d{3}Z$')
    
    excel_files = []
    
    for item in base_path.iterdir():
        if item.is_dir() and timestamp_pattern.match(item.name):
            xlsx_file = item / "htsdata.xlsx"
            if xlsx_file.exists():
                excel_files.append({
                    'timestamp_dir': item.name,
                    'xlsx_file': xlsx_file,
                    'csv_file': item / "htsdata.csv"  # For chapter identification
                })
    
    return excel_files

def identify_chapter_from_excel(excel_file):
    """Identify chapter directly from Excel file."""
    if not OPENPYXL_AVAILABLE:
        return None
    
    try:
        workbook = load_workbook(excel_file, read_only=True)
        worksheet = workbook.active
        
        # Get first row with data (skip header if exists)
        for row in worksheet.iter_rows(min_row=2, max_row=3, values_only=True):
            if row and row[0]:  # First column should be HTS Number
                hts_number = str(row[0]).strip('"').strip()
                
                if len(hts_number) >= 2:
                    try:
                        chapter_num = int(hts_number[:2])
                        if 1 <= chapter_num <= 99:
                            workbook.close()
                            return chapter_num
                    except ValueError:
                        continue
        
        workbook.close()
    except Exception as e:
        print(f"Error reading Excel file {excel_file}: {e}")
    
    return None

def identify_chapter_from_csv(csv_file):
    """Identify chapter from corresponding CSV file (fallback method)."""
    if not csv_file.exists():
        return None
    
    try:
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader)
            first_row = next(reader)
            hts_number = first_row[0].strip('"').strip()
            
            if len(hts_number) >= 2:
                chapter_num = int(hts_number[:2])
                if 1 <= chapter_num <= 99:
                    return chapter_num
                    
    except Exception as e:
        print(f"Error reading {csv_file}: {e}")
    
    return None

def organize_excel_files(playwright_dir, chapters_dir):
    """Organize Excel files by chapter."""
    print("🔍 Finding Excel files...")
    excel_files = find_excel_files(playwright_dir)
    print(f"Found {len(excel_files)} Excel files")
    
    # Group files by chapter
    chapter_files = {}
    unidentified_files = []
    
    print("📊 Analyzing files to identify chapters...")
    for file_info in excel_files:
        # Try to identify from Excel first, then fall back to CSV
        chapter_num = identify_chapter_from_excel(file_info['xlsx_file'])
        if not chapter_num:
            chapter_num = identify_chapter_from_csv(file_info['csv_file'])
        
        if chapter_num:
            chapter_name = CHAPTER_NAMES.get(chapter_num, f"chapter-{chapter_num}")
            print(f"  Chapter {chapter_num}: {chapter_name} ({file_info['timestamp_dir']})")
            
            # Only keep one Excel file per chapter (first one found)
            if chapter_num not in chapter_files:
                chapter_files[chapter_num] = file_info
            else:
                print(f"    (Duplicate found, keeping first one)")
        else:
            print(f"  Could not identify chapter for {file_info['timestamp_dir']}")
            unidentified_files.append(file_info)
    
    # Copy Excel files to chapter directories
    print("📁 Copying Excel files to chapter directories...")
    successful_copies = 0
    failed_copies = []
    
    for chapter_num, file_info in chapter_files.items():
        chapter_name = CHAPTER_NAMES.get(chapter_num, f"chapter-{chapter_num}")
        chapter_dir = Path(chapters_dir) / f"chapter-{chapter_num:02d}"
        
        # Create chapter directory if it doesn't exist
        chapter_dir.mkdir(parents=True, exist_ok=True)
        
        target_filename = f"chapter-{chapter_num:02d}-{chapter_name}.xlsx"
        target_path = chapter_dir / target_filename
        
        try:
            shutil.copy2(file_info['xlsx_file'], target_path)
            print(f"  ✅ Chapter {chapter_num:02d}: {target_filename}")
            successful_copies += 1
        except Exception as e:
            error_msg = f"Failed to copy {file_info['xlsx_file']} to {target_path}: {e}"
            print(f"  ❌ Chapter {chapter_num:02d}: {error_msg}")
            failed_copies.append(error_msg)
    
    return chapter_files, unidentified_files, successful_copies, failed_copies

def generate_excel_report(chapter_files, unidentified_files, successful_copies, failed_copies, output_dir):
    """Generate Excel organization report."""
    report_path = Path(output_dir) / "analysis" / "excel_organization_report.md"
    report_path.parent.mkdir(exist_ok=True)
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("# Excel Files Organization Report\n\n")
        
        f.write("## Summary\n\n")
        f.write(f"- **Excel Files Found**: {len(chapter_files) + len(unidentified_files)}\n")
        f.write(f"- **Chapters Identified**: {len(chapter_files)}\n")
        f.write(f"- **Unidentified Files**: {len(unidentified_files)}\n")
        f.write(f"- **Successful Copies**: {successful_copies}\n")
        f.write(f"- **Failed Copies**: {len(failed_copies)}\n\n")
        
        f.write("## Organized Chapters\n\n")
        f.write("| Chapter | Name | Source Directory |\n")
        f.write("|---------|------|------------------|\n")
        
        for chapter_num in sorted(chapter_files.keys()):
            chapter_name = CHAPTER_NAMES.get(chapter_num, f"chapter-{chapter_num}")
            source_dir = chapter_files[chapter_num]['timestamp_dir']
            f.write(f"| {chapter_num:02d} | {chapter_name} | {source_dir} |\n")
        
        if unidentified_files:
            f.write(f"\n## Unidentified Files\n\n")
            for file_info in unidentified_files:
                f.write(f"- {file_info['timestamp_dir']}\n")
        
        if failed_copies:
            f.write(f"\n## Failed Copies\n\n")
            for error in failed_copies:
                f.write(f"- {error}\n")
    
    print(f"📋 Excel report saved to {report_path}")

def main():
    """Main function."""
    playwright_dir = "../playwright-mcp-output11"
    chapters_dir = "../chapters"
    output_dir = "../"
    
    print("📊 Excel Files Organization")
    print("=" * 40)
    
    chapter_files, unidentified_files, successful_copies, failed_copies = organize_excel_files(
        playwright_dir, chapters_dir
    )
    
    generate_excel_report(
        chapter_files, unidentified_files, successful_copies, failed_copies, output_dir
    )
    
    # Final summary
    print(f"\n🎉 Excel Organization Complete!")
    print(f"📊 {len(chapter_files)} chapters with Excel files")
    print(f"✅ {successful_copies} Excel files successfully copied")
    if failed_copies:
        print(f"⚠️  {len(failed_copies)} copy operations failed")
    if unidentified_files:
        print(f"❓ {len(unidentified_files)} files could not be identified")

if __name__ == "__main__":
    main()
